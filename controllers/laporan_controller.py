import requests
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.borders import Border, Side
from database.connection import Database

class LaporanController:
    def __init__(self):
        self.conn = Database()

    def export_excel(self, data, headers, filename):
        if isinstance(data[0], dict):
            df = pd.DataFrame(data)
        elif isinstance(data[0], (list, tuple)):
            df = pd.DataFrame(data, columns=headers)
        else:
            raise ValueError("Format data tidak dikenali.")

        for col in df.columns:
            if "tanggal" in col.lower():
                df[col] = df[col].astype(str)

        df.to_excel(filename, index=False, startrow=5)

        wb = load_workbook(filename)
        ws = wb.active

        max_col = len(df.columns)
        max_row = len(df.index) + 1
        last_col_letter = get_column_letter(max_col)

        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = "KOST VIBE HOUSE"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"] = "Jl. Raya Buah Batu, Kota Bandung"
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells(f"A3:{last_col_letter}3")
        ws["A3"] = "0823-2030-8986"
        ws["A3"].alignment = Alignment(horizontal="center")

        header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        table_range = f"A6:{last_col_letter}{5 + max_row}"
        excel_table = Table(displayName="DataLaporan", ref=table_range)

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        excel_table.tableStyleInfo = style
        ws.add_table(excel_table)

        total_row = 5 + max_row + 1

        kolom_uang = None
        possible_fields = ["jumlah", "total_harga", "jumlah_bayar"]

        for i, col_name in enumerate(df.columns):
            if any(field in col_name.lower() for field in possible_fields):
                kolom_uang = get_column_letter(i + 1)
                break

        if kolom_uang:
            uang_col_index = ws[kolom_uang + "6"].column  # index kolom uang
            total_col_index = max_col  # kolom terakhir
            total_col_letter = get_column_letter(total_col_index)
            merge_end_letter = get_column_letter(total_col_index - 1)

            # Merge dari A sampai kolom sebelum kolom terakhir
            ws.merge_cells(f"A{total_row}:{merge_end_letter}{total_row}")
            ws[f"A{total_row}"] = "TOTAL PEMASUKAN"
            ws[f"A{total_row}"].font = Font(bold=True)
            ws[f"A{total_row}"].alignment = Alignment(horizontal="center")

            data_start_row = 7
            data_end_row = 5 + max_row

            # Hasil total ditaruh di kolom terakhir
            ws[f"{total_col_letter}{total_row}"] = f"=SUM({kolom_uang}{data_start_row}:{kolom_uang}{data_end_row})"
            ws[f"{total_col_letter}{total_row}"].font = Font(bold=True)
            ws[f"{total_col_letter}{total_row}"].alignment = Alignment(horizontal="center")
            ws[f"{total_col_letter}{total_row}"].number_format = '#,##0'

            # Styling
            total_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

            # Warnai seluruh baris TOTAL
            for col in range(1, total_col_index + 1):
                col_letter = get_column_letter(col)
                ws[f"{col_letter}{total_row}"].fill = total_fill
                ws[f"{col_letter}{total_row}"].border = border

        wb.save(filename)

    def get_laporan_pengeluaran(self, bulan: str, tahun: str, kategori: str = "Semua"):
        cursor = self.conn
        bulan_angka = self._bulan_ke_angka(bulan)
        print("DEBUG BULAN:", bulan, "->", bulan_angka)
        print("DEBUG TAHUN:", tahun)
        query = """
            SELECT kd_pengeluaran, tanggal, kategori, deskripsi, jumlah_harga, dibuat_oleh
            FROM pengeluaran
            WHERE MONTH(tanggal) = %s AND YEAR(tanggal) = %s
        """
        params = [self._bulan_ke_angka(bulan), tahun]

        if kategori != "Semua":
            query += " AND kategori = %s"
            params.append(kategori)

        cursor.execute(query, params)
        return cursor.fetchall()

    def get_laporan_bulanan(self, bulan: str, tahun: str, kategori: str = "Semua"):
        bulan_angka = self._bulan_ke_angka(bulan)

        if kategori == "Pengeluaran":
            return self.get_laporan_pengeluaran(bulan, tahun)

        elif kategori == "Pemasukan" or kategori == "Semua":
            cursor = self.conn
            query = """
                SELECT 
                    t.kd_transaksi,
                    p.nama AS nama_penyewa,
                    t.kd_unit,
                    t.tanggal_transaksi,
                    t.total_harga,
                    t.diskon,
                    t.biaya_tambahan,
                    t.jumlah_bayar,
                    t.uang_penyewa,
                    t.kembalian,
                    t.status_transaksi
                FROM transaksi t
                JOIN transaksi_bulanan tb ON t.kd_transaksi_bulanan = tb.kd_transaksi_bulanan
                JOIN penyewa p ON t.kd_penyewa = p.kd_penyewa
                WHERE tb.nama_bulan = %s AND tb.tahun = %s
            """
            params = [bulan, tahun]
            cursor.execute(query, params)
            return cursor.fetchall()

        else:
            return []

    def _bulan_ke_angka(self, nama_bulan):
        daftar_bulan = {
            "Januari": 1, "Februari": 2, "Maret": 3, "April": 4,
            "Mei": 5, "Juni": 6, "Juli": 7, "Agustus": 8,
            "September": 9, "Oktober": 10, "November": 11, "Desember": 12
        }
        return daftar_bulan.get(nama_bulan, 0)

    def get_pemasukan_periode(self, tgl_dari, tgl_sampai):
        query = """
            SELECT 
                t.kd_transaksi,
                p.nama AS nama_penyewa,
                t.kd_unit,
                t.tanggal_transaksi,
                t.total_harga,
                t.diskon,
                t.biaya_tambahan,
                t.jumlah_bayar,
                t.uang_penyewa,
                t.kembalian,
                t.status_transaksi
            FROM transaksi t
            JOIN transaksi_bulanan tb ON t.kd_transaksi_bulanan = tb.kd_transaksi_bulanan
            JOIN penyewa p ON t.kd_penyewa = p.kd_penyewa
            WHERE tanggal_transaksi BETWEEN %s AND %s
            ORDER BY tanggal_transaksi ASC
        """
        self.conn.execute(query, (tgl_dari, tgl_sampai))
        return self.conn.fetchall()

    def get_pengeluaran_periode(self, tgl_dari, tgl_sampai):
        query = """
            SELECT kd_pengeluaran, tanggal, kategori, deskripsi, jumlah_harga, dibuat_oleh
            FROM pengeluaran
            WHERE tanggal BETWEEN %s AND %s
            ORDER BY tanggal ASC
        """
        self.conn.execute(query, (tgl_dari, tgl_sampai))
        return self.conn.fetchall()

    def get_nomor_admin(self):
        cursor = self.conn
        query = "SELECT no_hp FROM users WHERE role = 'admin' LIMIT 1"
        cursor.execute(query)
        result = cursor.fetchone()
        return result['no_hp'] if result else None

    def kirim_wa_admin(self, message: str, file_path: str = None):
        cursor = self.conn
        cursor.execute("SELECT no_hp FROM users WHERE role = 'admin'")
        admin = cursor.fetchone()

        if admin and admin['no_hp']:
            no_hp = admin['no_hp']
            token = "FxOEioHDnpQ3liM6zjIRcNlSGSVh1dF80jCY738OrubljoZovkRProM"
            headers = {
                "Authorization": token
            }
            if file_path and os.path.exists(file_path):
                url = "https://sby.wablas.com/api/v2/send-document"
                files = {
                    "document": open(file_path, "rb")
                }
                data = {
                    "phone": no_hp,
                    "caption": message
                }
                response = requests.post(url, headers=headers, files=files, data=data)
            else:
                url = "https://sby.wablas.com/api/send-message"
                headers["Content-Type"] = "application/json"
                payload = {
                    "phone": no_hp,
                    "message": message
                }
                response = requests.post(url, headers=headers, json=payload)
            print("WA Admin Status:", response.status_code, response.text)